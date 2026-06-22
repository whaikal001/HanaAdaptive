import os
import io
import csv
import json
import sqlite3
import secrets
from functools import wraps

from flask import (
    Flask, request, session, redirect, url_for,
    render_template, jsonify, Response, abort, g,
)

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "hana-admin")
INGEST_TOKEN = os.environ.get("INGEST_TOKEN", "hana-secret-token")
SECRET_KEY = os.environ.get("SECRET_KEY", secrets.token_hex(32))
DB_PATH = os.environ.get("HANA_DB_PATH", os.path.join(os.path.dirname(__file__), "hana.db"))

app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY
INTERACTION_FIELDS = [
    "client_ts", "question", "answer", "score", "stress_level",
    "Ea", "Ec", "Ad", "Be", "D", "Ep", "empathy_mode",
]

SUMMARY_FIELDS = [
    "client_ts", "start_stress_score", "end_stress_score", "stress_improvement",
    "final_calm_rating", "modes_used", "techniques_used", "total_interactions",
    "session_duration_minutes", "end_action",
    "stress_level", "empathy_mode", "intention_level", "empathy_E",
    "support_need_announced",
]


def get_db():
    if "db" not in g:
        # timeout: wait up to 30s for a competing writer instead of hanging the
        # (single, on PythonAnywhere) worker forever — SQLite locking over the
        # host's network filesystem can stall under the client's retry storm.
        g.db = sqlite3.connect(DB_PATH, timeout=30)
        g.db.row_factory = sqlite3.Row
        # WAL lets the dashboard read while the game is writing, so reads don't
        # block writes (and vice-versa).
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA busy_timeout=30000")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH, timeout=30)
    db.execute("PRAGMA journal_mode=WAL")
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS interactions (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            received_at     TEXT DEFAULT (datetime('now')),
            device_id       TEXT,
            app_type        TEXT,
            user_name       TEXT,
            session_id      TEXT,
            session_file    TEXT,
            client_ts       TEXT,
            question        TEXT,
            answer          TEXT,
            score           TEXT,
            stress_level    TEXT,
            Ea              REAL,
            Ec              REAL,
            Ad              REAL,
            Be              REAL,
            D               REAL,
            Ep              REAL,
            empathy_mode    TEXT,
            extra           TEXT
        );

        CREATE TABLE IF NOT EXISTS summaries (
            id                       INTEGER PRIMARY KEY AUTOINCREMENT,
            received_at              TEXT DEFAULT (datetime('now')),
            device_id                TEXT,
            app_type                 TEXT,
            user_name                TEXT,
            session_id               TEXT,
            session_file             TEXT,
            client_ts                TEXT,
            start_stress_score       TEXT,
            end_stress_score         TEXT,
            stress_improvement       TEXT,
            final_calm_rating        TEXT,
            modes_used               TEXT,
            techniques_used          TEXT,
            total_interactions       TEXT,
            session_duration_minutes TEXT,
            end_action               TEXT,
            stress_level             TEXT,
            empathy_mode             TEXT,
            intention_level          TEXT,
            empathy_E                TEXT,
            support_need_announced   TEXT,
            attempts                 INTEGER DEFAULT 1
        );

        CREATE INDEX IF NOT EXISTS idx_inter_device  ON interactions(device_id);
        CREATE INDEX IF NOT EXISTS idx_inter_session ON interactions(session_id);
        CREATE INDEX IF NOT EXISTS idx_summ_device   ON summaries(device_id);
        """
    )

    # Lightweight migration: CREATE TABLE IF NOT EXISTS won't add new columns to a
    # summaries table that already exists on a deployed DB, so add any that are
    # missing. New columns default to NULL for older rows.
    existing_cols = {row[1] for row in db.execute("PRAGMA table_info(summaries)")}
    for col in ("stress_level", "empathy_mode", "intention_level",
                "empathy_E", "support_need_announced", "app_type",
                "profile_slot", "profile_id"):
        if col not in existing_cols:
            db.execute("ALTER TABLE summaries ADD COLUMN %s TEXT" % col)
    if "attempts" not in existing_cols:
        db.execute("ALTER TABLE summaries ADD COLUMN attempts INTEGER DEFAULT 1")

    # Same for interactions: tag each row with the study arm + profile id/slot.
    inter_cols = {row[1] for row in db.execute("PRAGMA table_info(interactions)")}
    if "app_type" not in inter_cols:
        db.execute("ALTER TABLE interactions ADD COLUMN app_type TEXT")
    if "profile_slot" not in inter_cols:
        db.execute("ALTER TABLE interactions ADD COLUMN profile_slot TEXT")
    if "profile_id" not in inter_cols:
        db.execute("ALTER TABLE interactions ADD COLUMN profile_id TEXT")

    db.commit()
    db.close()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def check_ingest_token(req):
    token = req.headers.get("X-Ingest-Token", "")
    return secrets.compare_digest(token, INGEST_TOKEN)


@app.route("/api/interaction", methods=["POST"])
def api_interaction():
    if not check_ingest_token(request):
        return jsonify(ok=False, error="bad token"), 401

    data = request.get_json(silent=True) or {}

    # Anything outside the known columns is stashed in `extra` so we never lose
    # the enhanced-logging fields (Ea_history, struggle metrics, etc.).
    known = {"device_id", "app_type", "user_name", "profile_id", "profile_slot", "session_id", "session_file"} | set(INTERACTION_FIELDS)
    extra = {k: v for k, v in data.items() if k not in known}

    db = get_db()
    db.execute(
        """
        INSERT INTO interactions
            (device_id, app_type, user_name, profile_id, profile_slot, session_id, session_file, client_ts,
             question, answer, score, stress_level,
             Ea, Ec, Ad, Be, D, Ep, empathy_mode, extra)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            data.get("device_id"), data.get("app_type"),
            data.get("user_name"), _s(data.get("profile_id")), _s(data.get("profile_slot")), data.get("session_id"),
            data.get("session_file"), data.get("client_ts"),
            data.get("question"), data.get("answer"), _s(data.get("score")),
            data.get("stress_level"),
            _f(data.get("Ea")), _f(data.get("Ec")), _f(data.get("Ad")),
            _f(data.get("Be")), _f(data.get("D")), _f(data.get("Ep")),
            data.get("empathy_mode"),
            json.dumps(extra) if extra else None,
        ),
    )
    db.commit()
    return jsonify(ok=True)


@app.route("/api/summary", methods=["POST"])
def api_summary():
    if not check_ingest_token(request):
        return jsonify(ok=False, error="bad token"), 401

    data = request.get_json(silent=True) or {}
    db = get_db()

    # Each profile gets a stable uuid (profile_id) when it's created in a slot;
    # that's the per-profile identity we dedup on. profile_slot is kept too, but
    # only as informational context (it gets reused when a slot is reset).
    profile_id = data.get("profile_id")
    profile_id = str(profile_id) if profile_id is not None and str(profile_id).strip() != "" else None
    slot = data.get("profile_slot")
    slot = str(slot) if slot is not None and str(slot).strip() != "" else None

    fields = (
        profile_id, slot,
        data.get("session_id"), data.get("session_file"), data.get("client_ts"),
        _s(data.get("start_stress_score")), _s(data.get("end_stress_score")),
        _s(data.get("stress_improvement")), _s(data.get("final_calm_rating")),
        data.get("modes_used"), data.get("techniques_used"),
        _s(data.get("total_interactions")), _s(data.get("session_duration_minutes")),
        data.get("end_action"),
        data.get("stress_level"), data.get("empathy_mode"),
        data.get("intention_level"), _s(data.get("empathy_E")),
        _s(data.get("support_need_announced")),
    )

    # One row per profile: replaying the SAME profile overwrites its summary
    # (attempts++), while a DIFFERENT profile — even same username, or a new
    # profile created after a slot was reset — gets a fresh profile_id and its
    # own row, so reset+reuse never clobbers earlier data. Fall back to session_id
    # for builds that don't send a profile_id yet (those never overwrite either).
    if profile_id is not None:
        existing = db.execute(
            "SELECT id, attempts FROM summaries WHERE device_id = ? AND profile_id = ?",
            (data.get("device_id"), profile_id),
        ).fetchone()
    else:
        existing = db.execute(
            "SELECT id, attempts FROM summaries WHERE device_id = ? AND session_id = ?",
            (data.get("device_id"), data.get("session_id")),
        ).fetchone()

    if existing:
        db.execute(
            """
            UPDATE summaries SET
                received_at = datetime('now'),
                app_type = ?,
                profile_id = ?, profile_slot = ?,
                session_id = ?, session_file = ?, client_ts = ?,
                start_stress_score = ?, end_stress_score = ?, stress_improvement = ?,
                final_calm_rating = ?, modes_used = ?, techniques_used = ?,
                total_interactions = ?, session_duration_minutes = ?, end_action = ?,
                stress_level = ?, empathy_mode = ?, intention_level = ?, empathy_E = ?,
                support_need_announced = ?,
                attempts = ?
            WHERE id = ?
            """,
            (data.get("app_type"),) + fields + ((existing["attempts"] or 1) + 1, existing["id"]),
        )
    else:
        db.execute(
            """
            INSERT INTO summaries
                (device_id, app_type, user_name, profile_id, profile_slot, session_id, session_file,
                 client_ts, start_stress_score, end_stress_score, stress_improvement,
                 final_calm_rating, modes_used, techniques_used,
                 total_interactions, session_duration_minutes, end_action,
                 stress_level, empathy_mode, intention_level, empathy_E,
                 support_need_announced, attempts)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
            """,
            (data.get("device_id"), data.get("app_type"), data.get("user_name")) + fields,
        )
    db.commit()
    return jsonify(ok=True)


@app.route("/api/health")
def health():
    return jsonify(ok=True)


def _f(v):
    """Best-effort float; keeps NULLs as NULL."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _s(v):
    """Store scalars as text without turning None into the string 'None'."""
    if v is None:
        return None
    return str(v)

@app.route("/")
def index():
    return redirect(url_for("dashboard"))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if secrets.compare_digest(request.form.get("password", ""), ADMIN_PASSWORD):
            session["admin"] = True
            return redirect(request.args.get("next") or url_for("dashboard"))
        error = "Wrong password."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    db = get_db()
    users = db.execute(
        """
        SELECT
            device_id,
            MAX(user_name)                          AS user_name,
            COUNT(*)                                AS interactions,
            COUNT(DISTINCT session_id)              AS sessions,
            MAX(received_at)                        AS last_seen
        FROM interactions
        GROUP BY device_id
        ORDER BY last_seen DESC
        """
    ).fetchall()

    totals = db.execute(
        "SELECT COUNT(*) AS i, COUNT(DISTINCT device_id) AS d, "
        "COUNT(DISTINCT session_id) AS s FROM interactions"
    ).fetchone()
    summ_total = db.execute("SELECT COUNT(*) AS c FROM summaries").fetchone()["c"]

    return render_template(
        "dashboard.html", users=users, totals=totals, summ_total=summ_total
    )


@app.route("/user/<device_id>")
@login_required
def user_detail(device_id):
    db = get_db()
    sessions = db.execute(
        """
        SELECT
            session_id,
            MAX(user_name)  AS user_name,
            COUNT(*)        AS turns,
            MIN(client_ts)  AS started,
            MAX(client_ts)  AS ended
        FROM interactions
        WHERE device_id = ?
        GROUP BY session_id
        ORDER BY started DESC
        """,
        (device_id,),
    ).fetchall()

    summaries = db.execute(
        "SELECT * FROM summaries WHERE device_id = ? ORDER BY received_at DESC",
        (device_id,),
    ).fetchall()

    name = sessions[0]["user_name"] if sessions else device_id
    return render_template(
        "user.html", device_id=device_id, name=name,
        sessions=sessions, summaries=summaries,
    )


@app.route("/session/<device_id>/<session_id>")
@login_required
def session_detail(device_id, session_id):
    db = get_db()
    rows = db.execute(
        """
        SELECT * FROM interactions
        WHERE device_id = ? AND session_id = ?
        ORDER BY id ASC
        """,
        (device_id, session_id),
    ).fetchall()
    name = rows[0]["user_name"] if rows else ""
    return render_template(
        "session.html", rows=rows, device_id=device_id,
        session_id=session_id, name=name,
    )


# The three study arms, as stored in `app_type`, with display labels.
ARM_LABELS = [
    ("adaptive", "Adaptive"),
    ("strict_empathy", "Fixed (strict empathy)"),
    ("neutral", "Neutral (baseline)"),
]
ARM_SHORT = {"adaptive": "Adaptive", "strict_empathy": "Fixed", "neutral": "Neutral"}


@app.route("/summaries")
@login_required
def summaries():
    db = get_db()

    # Optional ?arm=<adaptive|strict_empathy|neutral> filter — "all" (or missing)
    # shows everything.
    arm = request.args.get("arm", "all")
    valid_arms = {k for k, _ in ARM_LABELS}
    # Order by the device timestamp (when the session actually happened), falling
    # back to server-receive time for rows that predate client_ts. client_ts is
    # "%Y-%m-%d %H:%M:%S" text, so it sorts chronologically as a string.
    if arm in valid_arms:
        rows = db.execute(
            "SELECT * FROM summaries WHERE app_type = ? "
            "ORDER BY COALESCE(client_ts, received_at) DESC",
            (arm,),
        ).fetchall()
    else:
        arm = "all"
        rows = db.execute(
            "SELECT * FROM summaries "
            "ORDER BY COALESCE(client_ts, received_at) DESC"
        ).fetchall()

    # Per-arm counts for the filter tabs / stat cards.
    counts = {k: 0 for k, _ in ARM_LABELS}
    total = 0
    for r in db.execute("SELECT app_type, COUNT(*) AS c FROM summaries GROUP BY app_type"):
        total += r["c"]
        if r["app_type"] in counts:
            counts[r["app_type"]] = r["c"]

    return render_template(
        "summaries.html", rows=rows, arm=arm,
        arm_labels=ARM_LABELS, arm_short=ARM_SHORT, counts=counts, total=total,
    )


@app.route("/summary/<int:summary_id>/delete", methods=["POST"])
@login_required
def delete_summary(summary_id):
    db = get_db()
    db.execute("DELETE FROM summaries WHERE id = ?", (summary_id,))
    db.commit()
    return redirect(request.referrer or url_for("summaries"))


@app.route("/summaries/delete", methods=["POST"])
@login_required
def delete_summaries_bulk():
    """Delete every summary in the current view: one arm, or all of them."""
    arm = request.form.get("arm", "all")
    db = get_db()
    if arm in {k for k, _ in ARM_LABELS}:
        db.execute("DELETE FROM summaries WHERE app_type = ?", (arm,))
        db.commit()
        return redirect(url_for("summaries", arm=arm))
    db.execute("DELETE FROM summaries")
    db.commit()
    return redirect(url_for("summaries"))


@app.route("/user/<device_id>/delete", methods=["POST"])
@login_required
def delete_user(device_id):
    """Remove a device/user entirely: all their interactions and summaries."""
    db = get_db()
    db.execute("DELETE FROM interactions WHERE device_id = ?", (device_id,))
    db.execute("DELETE FROM summaries WHERE device_id = ?", (device_id,))
    db.commit()
    return redirect(url_for("dashboard"))


@app.route("/export/interactions.csv")
@login_required
def export_interactions():
    return _export("interactions")


@app.route("/export/summaries.csv")
@login_required
def export_summaries():
    return _export("summaries")


# One Excel workbook with a separate sheet per study arm, so all three HANA
# versions land in a single file: "adaptive", "strict_empathy", "baseline".
ARM_SHEET_ORDER = ["adaptive", "strict_empathy", "baseline"]


@app.route("/export/summaries.xlsx")
@login_required
def export_summaries_xlsx():
    return _export_xlsx("summaries")


@app.route("/export/interactions.xlsx")
@login_required
def export_interactions_xlsx():
    return _export_xlsx("interactions")


def _export_xlsx(table):
    from openpyxl import Workbook

    db = get_db()
    rows = db.execute("SELECT * FROM %s ORDER BY id ASC" % table).fetchall()
    cols = list(rows[0].keys()) if rows else []

    # Group every row onto a sheet named after its study arm. Rows with no
    # app_type (e.g. data from before this column existed) go to "untagged".
    grouped = {}
    for r in rows:
        arm = (r["app_type"] if "app_type" in r.keys() else None) or "untagged"
        grouped.setdefault(arm, []).append(r)

    sheet_order = list(ARM_SHEET_ORDER)
    for arm in grouped:
        if arm not in sheet_order:
            sheet_order.append(arm)

    wb = Workbook()
    wb.remove(wb.active)
    for arm in sheet_order:
        ws = wb.create_sheet(title=arm[:31])  # Excel caps sheet names at 31 chars
        if cols:
            ws.append(cols)
        for r in grouped.get(arm, []):
            ws.append([r[k] for k in cols])
    if not wb.sheetnames:
        wb.create_sheet(title="empty")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        bio.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=hana_%s.xlsx" % table},
    )


def _export(table):
    db = get_db()
    rows = db.execute("SELECT * FROM %s ORDER BY id ASC" % table).fetchall()
    out = io.StringIO()
    writer = csv.writer(out)
    if rows:
        writer.writerow(rows[0].keys())
        for r in rows:
            writer.writerow([r[k] for k in r.keys()])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=%s.csv" % table},
    )


# Initialise the schema as soon as the module is imported, so it works both
# under `flask run`, `python app.py`, and WSGI hosting (gunicorn / PythonAnywhere).
init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
